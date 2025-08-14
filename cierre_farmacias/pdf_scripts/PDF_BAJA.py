from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import win32com.client
import os

# Nota: Este script conserva la lógica original pero ahora vive como módulo de la app.
# Lee la cadena de conexión desde variables internas del propio archivo (puedes migrarlo a config si se desea).

import os
DB_SERVER = os.getenv('DB_SERVER', 'localhost')
DB_NAME = os.getenv('DB_NAME', 'DBBI')
DB_USER = os.getenv('DB_USER', 'sa')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_TABLE = 'CierreSucursales4'


def get_data_from_sql(departamento, ceco):
    SQLALCHEMY_DATABASE_URI = f"mssql+pyodbc://{DB_USER}:{DB_PASSWORD}@{DB_SERVER}/{DB_NAME}?driver=ODBC+Driver+17+for+SQL+Server"
    engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=False)
    query = text("""
    SELECT 
        [Act. Fijo], [Tipo de Activo], [Denominacion del activo fijo], [Val. Cont.],
        'NA' as Modelo, 'NA' as Serie, [Ceco],
        STUFF((SELECT ',' + [Nombre_del_Proyecto]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Nombre_del_Proyecto,
        STUFF((SELECT ',' + [Tipo_de_Proyecto]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Tipo_de_Proyecto,
        [DetallesFirmaSolicitante], [Departamento],
        STUFF((SELECT ',' + [Observaciones]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Observaciones,
        STUFF((SELECT ',' + [Detalle o relación del activo Fijo]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Detalle o relación del activo Fijo],
        STUFF((SELECT ',' + [Copia Factura de compra]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Copia Factura de compra],
        STUFF((SELECT ',' + [Correo de cierre de Sucursal]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Correo de cierre de Sucursal],
        STUFF((SELECT ',' + [Informe o Dictamen tecnico]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Informe o Dictamen tecnico],
        STUFF((SELECT ',' + [Dictamen Aseguradora]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Dictamen Aseguradora],
        STUFF((SELECT ',' + [Calculo de cobro para venta]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Calculo de cobro para venta]
    FROM [DBBI].[dbo].[CierreSucursales4] AS T1
    WHERE [Ceco] = :ceco AND [Departamento] = :departamento AND [Accion] = 'BAJA'
    GROUP BY [Act. Fijo], [Tipo de Activo], [Denominacion del activo fijo], [Val. Cont.], [Ceco],Departamento,[DetallesFirmaSolicitante]
    """)
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        return result.fetchall()


def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    ruta_temp = None
    excel = None
    try:
        datos = get_data_from_sql(departamento, ceco)
        if not datos:
            return
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        ruta_pdf = os.path.join(output_dir, f"Formato Baja_{departamento}_{ceco}.pdf")
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        wb = load_workbook(filename=ruta_excel)
        if 'BAJA' not in wb.sheetnames:
            return
        hoja_baja = wb['BAJA']
        for i, row in enumerate(datos):
            fila_actual = 32 + i
            hoja_baja[f'B{fila_actual}'] = row[0]
            hoja_baja[f'C{fila_actual}'] = row[1]
            hoja_baja[f'F{fila_actual}'] = row[2]
            hoja_baja[f'J{fila_actual}'] = row[3]
            hoja_baja[f'K{fila_actual}'] = row[4]
            hoja_baja[f'N{fila_actual}'] = row[5]
            hoja_baja[f'O{fila_actual}'] = row[6]
            hoja_baja['D17'] = row[7]
            hoja_baja['D18'] = row[8]
            hoja_baja['D19'] = row[9]
            hoja_baja['D20'] = row[10]
            hoja_baja['B26'] = row[11]
        wb.save(ruta_temp)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        hoja_baja = wb.Sheets("BAJA")
        hoja_baja.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
    finally:
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
