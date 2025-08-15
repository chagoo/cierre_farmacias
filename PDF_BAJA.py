from sqlalchemy import text
from openpyxl import load_workbook
import win32com.client
import os
import sys
from db_config import DB_TABLE, get_engine


def get_data_from_sql(departamento, ceco):
    """
    Obtiene los datos de la base de datos usando los parámetros departamento y ceco.
    """
    engine = get_engine()
    
    # query = text(f"""
    # SELECT [Act. Fijo],[Tipo de Activo],[Denominacion del activo fijo], [Val. Cont.], 'NA' as Modelo, 'NA' as Serie, [Ceco] 
    # FROM {DB_TABLE}
    # WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    # """)

    # query = text(f"""
    #  SELECT top 1 [Act. Fijo], [Tipo de Activo], [Denominacion del activo fijo], [Val. Cont.], 'NA' as Modelo, 'NA' as Serie, [Ceco],[Nombre_del_Proyecto], [Tipo_de_Proyecto],[DetallesFirmaSolicitante],[Departamento],
#    [Observaciones],    [Detalle o relación del activo Fijo]
    
    # FROM {DB_TABLE}
    # WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    # """)

    query = text("""
    SELECT 
        [Act. Fijo], 
        [Tipo de Activo], 
        [Denominacion del activo fijo], 
        [Val. Cont.], 
        'NA' as Modelo, 
        'NA' as Serie, 
        [Ceco],
        STUFF((SELECT ',' + [Nombre_del_Proyecto]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Nombre_del_Proyecto,
        STUFF((SELECT ',' + [Tipo_de_Proyecto]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Tipo_de_Proyecto,
                 
       
                                             
           [DetallesFirmaSolicitante],
		   [Departamento],
                 
        STUFF((SELECT ',' + [Observaciones]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS Observaciones,
        STUFF((SELECT ',' + [Detalle o relación del activo Fijo]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Detalle o relación del activo Fijo],
        STUFF((SELECT ',' + [Copia Factura de compra]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Copia Factura de compra],
        STUFF((SELECT ',' + [Correo de cierre de Sucursal]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Correo de cierre de Sucursal],
        STUFF((SELECT ',' + [Informe o Dictamen tecnico]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Informe o Dictamen tecnico],
        STUFF((SELECT ',' + [Dictamen Aseguradora]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Dictamen Aseguradora],
        STUFF((SELECT ',' + [Calculo de cobro para venta]
               FROM [DBBI].[dbo].[CierreSucursales4] AS T2
               WHERE T2.[Ceco] = T1.[Ceco] AND T2.[Departamento]=:departamento AND T2.[Accion] = 'BAJA'
               FOR XML PATH('')), 1, 1, '') AS [Calculo de cobro para venta]
    FROM 
        [DBBI].[dbo].[CierreSucursales4] AS T1
    WHERE 
        [Ceco] = :ceco AND [Departamento] = :departamento AND [Accion] = 'BAJA'
    GROUP BY 
        [Act. Fijo], 
        [Tipo de Activo], 
        [Denominacion del activo fijo], 
        [Val. Cont.], 
        [Ceco],Departamento,[DetallesFirmaSolicitante]
""")


  
    
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        return result.fetchall()

def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    """
    Actualiza el Excel con datos de SQL y genera PDF con nombre dinámico en directorio específico.
    """
    ruta_temp = None
    excel = None
    
    try:
        # Obtener datos de SQL
        datos = get_data_from_sql(departamento, ceco)
        if not datos:
            print(f"No se encontraron datos para Departamento: {departamento} y CECO: {ceco}")
            return
        
        print(f"Se encontraron {len(datos)} registros para procesar")
        
        # Crear directorio de salida si no existe
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"Carpeta creada: {output_dir}")
        
        # Generar ruta de PDF con nombre dinámico
        ruta_pdf = os.path.join(output_dir, f"Formato Baja_{departamento}_{ceco}.pdf")
        
        # Crear una copia temporal del template
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        
        # Cargar y modificar el Excel con openpyxl
        wb = load_workbook(filename=ruta_excel)
        if 'BAJA' not in wb.sheetnames:
            print("Error: No se encontró la hoja 'BAJA' en el archivo Excel")
            return
            
        hoja_baja = wb['BAJA']
        
        # Llenar datos en Excel
        for i, row in enumerate(datos):
            fila_actual = 32 + i  # Comenzamos en la fila 32 y avanzamos según el número de registro
            
            
            # Mapeo directo de campos SQL a celdas Excel
            hoja_baja[f'B{fila_actual}'] = row[0]  # ActFijo
            hoja_baja[f'C{fila_actual}'] = row[1]  # TipoActivo
            hoja_baja[f'F{fila_actual}'] = row[2]  # Title
            hoja_baja[f'J{fila_actual}'] = row[3]  # ValCont
            hoja_baja[f'K{fila_actual}'] = row[4]  # Modelo
            hoja_baja[f'N{fila_actual}'] = row[5]  # Serie
            hoja_baja[f'O{fila_actual}'] = row[6]  # Ceco
           

            hoja_baja['D17'] = row[7]              #[Nombre_del_Proyecto]
            hoja_baja['D18'] = row[8]              # [Tipo_de_Proyecto]
            hoja_baja['D19'] = row[9]              #[DetallesFirmaSolicitante]
            hoja_baja['D20'] = row[10]            #[Departamento],
            # hoja_baja['H24'] = row[7]
            hoja_baja['B26'] = row[11]    # [Observaciones],
             # Lógica para marcar "X" en las celdas correspondie
            valor_activo = row[12]
            celdas_x = {
               "Venta": ["F57", "D11"],
               "Cierre Sucursal": ["G57", "D13"],
               "Extravío o Robo": ["K57", "H11"],
               "Destrucción": ["J57", "H10"],
               "Siniestro": ["N57", "H12"],
               "Otros": ["O57", "H13"]

               
 }
            
            valor_activo2 = row[13]
            celdas_A = {
               "Venta": ["F58", "D11"],
               "Cierre Sucursal": ["G58", "D13"],
               "Extravío o Robo": ["K58", "H11"],
               "Destrucción": ["J58", "H10"],
               "Siniestro": ["N58", "H12"],
               "Otros": ["O58", "H13"]

               
 }
            

            valor_activo3 = row[14]
            celdas_B = {
               "Venta": ["F59", "D11"],
               "Cierre Sucursal": ["G59", "D13"],
               "Extravío o Robo": ["K59", "H11"],
               "Destrucción": ["J59", "H10"],
               "Siniestro": ["N59", "H12"],
               "Otros": ["O59", "H13"]

               
 }
 

            valor_activo4 = row[15]
            celdas_C = {
                "Venta": ["F60", "D11"],
                "Cierre Sucursal": ["G60", "D13"],
                "Extravío o Robo": ["K60", "H11"],
                "Destrucción": ["J60", "H10"],
                "Siniestro": ["L60", "H12"],
                "Otros": ["O60", "H13"]
              
  }
            

            valor_activo5 = row[16]
            celdas_D = {
                "Venta": ["F61", "D11"],
                "Cierre Sucursal": ["G61", "D13"],
                "Extravío o Robo": ["K61", "H11"],
                "Destrucción": ["J61", "H10"],
                "Siniestro": ["L61", "H12"],
                "Otros": ["O61", "H13"]
              
  }

            valor_activo6 = row[17]
            celdas_E = {
               "Venta": ["F62", "D11"],
               "Cierre Sucursal": ["G62", "D13"],
               "Extravío o Robo": ["K62", "H11"],
               "Destrucción": ["J62", "H10"],
               "Siniestro": ["L62", "H12"],
               "Otros": ["O62", "H13"]
             
 }
 
 
 
            if valor_activo in celdas_x:
               for celda in celdas_x[valor_activo]:
                   hoja_baja[celda] = "X"                                               
   
            if valor_activo2 in celdas_A:
              for celda in celdas_A[valor_activo2]:
                  hoja_baja[celda] = "X"     

            if valor_activo3 in celdas_B:
              for celda in celdas_B[valor_activo3]:
                  hoja_baja[celda] = "X"                          
                             
            if valor_activo4 in celdas_C:
              for celda in celdas_C[valor_activo4]:
                  hoja_baja[celda] = "X"  

            if valor_activo5 in celdas_D:
              for celda in celdas_D[valor_activo5]:
               hoja_baja[celda] = "X"            
            
            if valor_activo6 in celdas_E:
               for celda in celdas_E[valor_activo6]:
                hoja_baja[celda] = "X"            
            
            
                                           
            
            print(f"Registro {i+1} insertado en fila {fila_actual}")
          
            
        
        # Guardar temporalmente para generar PDF
        wb.save(ruta_temp)
        
        # Usar Excel para generar el PDF
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Abrir el archivo temporal
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        
        # Obtener la hoja específica
        hoja_baja = wb.Sheets("BAJA")
        
        # Exportar a PDF
        hoja_baja.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        
        # Cerrar Excel
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
        
        print(f"PDF generado exitosamente en: {ruta_pdf}")
        
    except Exception as e:
        print(f"Error durante el proceso: {str(e)}")
    finally:
        # Limpiar archivo temporal si existe
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
                print("Archivo temporal eliminado")
            except Exception as e:
                print(f"Error al eliminar archivo temporal: {str(e)}")
        
        # Asegurarse de que Excel se cierre
        if excel:
            try:
                excel.Quit()
            except:
                pass

if __name__ == "__main__":
    # Ruta base del template de Excel
    ruta_excel = r"P:\CierreFarmacias\Plantillas\Formatobajaactivo.xlsx"
    
    # Verificar argumentos de línea de comando
    if len(sys.argv) != 3:
        print("Uso: python script.py DEPARTAMENTO CECO")
        sys.exit(1)
        
    departamento = sys.argv[1]
    ceco = sys.argv[2]
    
    # Definir directorio de salida dinámicamente
    output_dir = fr"P:\UPLOAD\{departamento}\{ceco}"
    
    print(f"Procesando para Departamento: {departamento} y CECO: {ceco}")
    actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco)





















































































































