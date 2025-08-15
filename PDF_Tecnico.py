from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import win32com.client
import os
import sys

# Configuración de la base de datos
DB_SERVER = 'MPWPAS01'
DB_NAME = 'DBBI'
DB_USER = 'AlertDBBI'
DB_PASSWORD = 'P4$9'
DB_TABLE = 'CierreSucursales4'


def get_data_from_sql(departamento, ceco):
    """
    Obtiene los datos de la base de datos usando los parámetros departamento y ceco.
    """
    SQLALCHEMY_DATABASE_URI = f"mssql+pyodbc://{DB_USER}:{DB_PASSWORD}@{DB_SERVER}/{DB_NAME}?driver=ODBC+Driver+17+for+SQL+Server"
    engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=False)
    
    query = text(f"""
            
    SELECT FORMAT(SUM(CAST([Val. Cont.] AS DECIMAL(18,2))), 'C', 'es-MX') AS Total
    FROM {DB_TABLE}
    WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    """)
    
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        return result.fetchall()

def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    """
    Actualiza el Excel con datos de SQL y genera PDF con nombre dinámico en directorio específico.
    """
    ruta_temp = None
    excel = None
    
    try:
        # Obtener datos de SQL
        datos = get_data_from_sql(departamento, ceco)
        if not datos:
            print(f"No se encontraron datos para Departamento: {departamento} y CECO: {ceco}")
            return
        
        print(f"Se encontraron {len(datos)} registros para procesar")
        
        # Crear directorio de salida si no existe
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"Carpeta creada: {output_dir}")
        
        # Generar ruta de PDF con nombre dinámico
        ruta_pdf = os.path.join(output_dir, f"Informe Tecnico_{departamento}_{ceco}.pdf")
        
        # Crear una copia temporal del template
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        
        # Cargar y modificar el Excel con openpyxl
        wb = load_workbook(filename=ruta_excel)
        if 'Tecnico' not in wb.sheetnames:
            print("Error: No se encontró la hoja 'Tecnico' en el archivo Excel")
            return
            
        hoja_baja = wb['Tecnico']
        
        # Llenar datos en Excel
        for i, row in enumerate(datos):
            fila_actual = 14 + i  # Comenzamos en la fila 32 y avanzamos según el número de registro
            
            # Mapeo directo de campos SQL a celdas Excel
            hoja_baja[f'H{fila_actual}'] = row[0]  #  Title
            # hoja_baja[f'C{fila_actual}'] = row[1]  # TipoActivo
            # hoja_baja[f'F{fila_actual}'] = row[2]  # ActFijo
            # hoja_baja[f'J{fila_actual}'] = row[3]  # ValCont
            # hoja_baja[f'K{fila_actual}'] = row[4]  # Modelo
            # hoja_baja[f'N{fila_actual}'] = row[5]  # Serie
            # hoja_baja[f'O{fila_actual}'] = row[6]  # Ceco
            
            print(f"Registro {i+1} insertado en fila {fila_actual}")
        
        # Guardar temporalmente para generar PDF
        wb.save(ruta_temp)
        
        # Usar Excel para generar el PDF
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Abrir el archivo temporal
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        
        # Obtener la hoja específica
        hoja_baja = wb.Sheets("Tecnico")
        
        # Exportar a PDF
        hoja_baja.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        
        # Cerrar Excel
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
        
        print(f"PDF generado exitosamente en: {ruta_pdf}")
        
    except Exception as e:
        print(f"Error durante el proceso: {str(e)}")
    finally:
        # Limpiar archivo temporal si existe
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
                print("Archivo temporal eliminado")
            except Exception as e:
                print(f"Error al eliminar archivo temporal: {str(e)}")
        
        # Asegurarse de que Excel se cierre
        if excel:
            try:
                excel.Quit()
            except:
                pass

if __name__ == "__main__":
    # Ruta base del template de Excel
    ruta_excel = r"P:\CierreFarmacias\Plantillas\Informe Tecnico.xlsx"
    
    # Verificar argumentos de línea de comando
    if len(sys.argv) != 3:
        print("Uso: python script.py DEPARTAMENTO CECO")
        sys.exit(1)
        
    departamento = sys.argv[1]
    ceco = sys.argv[2]
    
    # Definir directorio de salida dinámicamente
    output_dir = fr"P:\UPLOAD\{departamento}\{ceco}"
    
    print(f"Procesando para Departamento: {departamento} y CECO: {ceco}")
    actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco)





















































































































