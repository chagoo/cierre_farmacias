from sqlalchemy import text
from openpyxl import load_workbook
import win32com.client
import os
import sys
from db_config import DB_TABLE, get_engine


def get_data_from_sql(departamento, ceco):
    """
    Obtiene los datos de la base de datos usando los parámetros departamento y ceco.
    """
    engine = get_engine()
    
    # Consulta para obtener los datos de los activos
    query = text(f"""
    SELECT distinct [Denominacion del activo fijo],[Tipo de Activo],[Act. Fijo], [Val. Cont.], 'NA' as Modelo, 'NA' as Serie, [Ceco],[Farmacia],[Domicilio],[Estado]
    FROM {DB_TABLE}
    WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    """)
    
    # Consulta para obtener información de la sucursal/CECO
    sucursal_query = text(f"""
    SELECT distinct [Ceco], [Farmacia], [Domicilio], [Ciudad], [Estado]
    FROM {DB_TABLE}
     WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
   
    """)
    
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        activos = result.fetchall()
        
        # Obtener información de la sucursal
        sucursal_result = connection.execute(sucursal_query, {"ceco": ceco, "departamento": departamento})
        sucursal_info = sucursal_result.fetchone()
        
        return activos, sucursal_info

def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    """
    Actualiza el Excel con datos de SQL y genera PDF con nombre dinámico en directorio específico.
    """
    ruta_temp = None
    excel = None
    
    try:
        # Obtener datos de SQL
        activos, sucursal_info = get_data_from_sql(departamento, ceco)
        if not activos:
            print(f"No se encontraron datos para Departamento: {departamento} y CECO: {ceco}")
            return
        
        print(f"Se encontraron {len(activos)} registros para procesar")
        
        # Crear directorio de salida si no existe
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"Carpeta creada: {output_dir}")
        
        # Generar ruta de PDF con nombre dinámico
        ruta_pdf = os.path.join(output_dir, f"Formato Recoleccion_{departamento}_{ceco}.pdf")
        
        # Crear una copia temporal del template
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        
        # Cargar y modificar el Excel con openpyxl
        wb = load_workbook(filename=ruta_excel)
        if 'Recoleccion' not in wb.sheetnames:
            print("Error: No se encontró la hoja 'Recoleccion' en el archivo Excel")
            return
            
        hoja_recoleccion = wb['Recoleccion']
        
        # Llenar datos de la sucursal en las celdas correspondientes
        if sucursal_info:
            hoja_recoleccion['A28'] = sucursal_info[0]  # CECO
            hoja_recoleccion['B28'] = sucursal_info[1]  # Inmueble (Farmacia)
            hoja_recoleccion['C28'] = sucursal_info[2]  # Domicilio
            hoja_recoleccion['H28'] = sucursal_info[3]  # Ciudad
            hoja_recoleccion['I28'] = sucursal_info[4]  # Estado
            print("Información de sucursal insertada en fila 28")
        
        # Llenar datos de los activos en Excel
        for i, row in enumerate(activos):
            fila_actual = 11 + i  # Comenzamos en la fila 11 y avanzamos según el número de registro
            
            # Mapeo directo de campos SQL a celdas Excel
            hoja_recoleccion[f'B{fila_actual}'] = row[0]  # Denominacion del activo fijo
            # hoja_recoleccion[f'C{fila_actual}'] = row[1]  # Tipo de Activo
            # hoja_recoleccion[f'F{fila_actual}'] = row[2]  # Act. Fijo
            # hoja_recoleccion[f'J{fila_actual}'] = row[3]  # Val. Cont.
            # hoja_recoleccion[f'K{fila_actual}'] = row[4]  # Modelo
            
            print(f"Registro {i+1} insertado en fila {fila_actual}")
        
        # Guardar temporalmente para generar PDF
        wb.save(ruta_temp)
        
        # Usar Excel para generar el PDF
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Abrir el archivo temporal
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        
        # Obtener la hoja específica
        hoja_recoleccion = wb.Sheets("Recoleccion")
        
        # Exportar a PDF
        hoja_recoleccion.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        
        # Cerrar Excel
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
        
        print(f"PDF generado exitosamente en: {ruta_pdf}")
        
    except Exception as e:
        print(f"Error durante el proceso: {str(e)}")
    finally:
        # Limpiar archivo temporal si existe
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
                print("Archivo temporal eliminado")
            except Exception as e:
                print(f"Error al eliminar archivo temporal: {str(e)}")
        
        # Asegurarse de que Excel se cierre
        if excel:
            try:
                excel.Quit()
            except:
                pass

if __name__ == "__main__":
    # Ruta base del template de Excel
    ruta_excel = r"P:\CierreFarmacias\Plantillas\Compromiso_Recoleccion.xlsx"
    
    # Verificar argumentos de línea de comando
    if len(sys.argv) != 3:
        print("Uso: python script.py DEPARTAMENTO CECO")
        sys.exit(1)
        
    departamento = sys.argv[1]
    ceco = sys.argv[2]
    
    # Definir directorio de salida dinámicamente
    output_dir = fr"P:\UPLOAD\{departamento}\{ceco}"
    
    print(f"Procesando para Departamento: {departamento} y CECO: {ceco}")
    actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco)