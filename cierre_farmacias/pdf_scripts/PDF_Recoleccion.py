from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import win32com.client
import os

import os
DB_SERVER = os.getenv('DB_SERVER', 'localhost')
DB_NAME = os.getenv('DB_NAME', 'DBBI')
DB_USER = os.getenv('DB_USER', 'sa')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_TABLE = 'CierreSucursales4'


def get_data_from_sql(departamento, ceco):
    SQLALCHEMY_DATABASE_URI = f"mssql+pyodbc://{DB_USER}:{DB_PASSWORD}@{DB_SERVER}/{DB_NAME}?driver=ODBC+Driver+17+for+SQL+Server"
    engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=False)
    query = text(f"""
    SELECT distinct [Denominacion del activo fijo],[Tipo de Activo],[Act. Fijo], [Val. Cont.], 'NA' as Modelo, 'NA' as Serie, [Ceco],[Farmacia],[Domicilio],[Estado]
    FROM {DB_TABLE}
    WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    """)
    sucursal_query = text(f"""
    SELECT distinct [Ceco], [Farmacia], [Domicilio], [Ciudad], [Estado]
    FROM {DB_TABLE}
     WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    """)
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        activos = result.fetchall()
        sucursal_result = connection.execute(sucursal_query, {"ceco": ceco, "departamento": departamento})
        sucursal_info = sucursal_result.fetchone()
        return activos, sucursal_info


def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    ruta_temp = None
    excel = None
    try:
        activos, sucursal_info = get_data_from_sql(departamento, ceco)
        if not activos:
            return
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        ruta_pdf = os.path.join(output_dir, f"Formato Recoleccion_{departamento}_{ceco}.pdf")
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        wb = load_workbook(filename=ruta_excel)
        if 'Recoleccion' not in wb.sheetnames:
            return
        hoja_recoleccion = wb['Recoleccion']
        if sucursal_info:
            hoja_recoleccion['A28'] = sucursal_info[0]
            hoja_recoleccion['B28'] = sucursal_info[1]
            hoja_recoleccion['C28'] = sucursal_info[2]
            hoja_recoleccion['H28'] = sucursal_info[3]
            hoja_recoleccion['I28'] = sucursal_info[4]
        for i, row in enumerate(activos):
            fila_actual = 11 + i
            hoja_recoleccion[f'B{fila_actual}'] = row[0]
        wb.save(ruta_temp)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        hoja_recoleccion = wb.Sheets("Recoleccion")
        hoja_recoleccion.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
    finally:
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
