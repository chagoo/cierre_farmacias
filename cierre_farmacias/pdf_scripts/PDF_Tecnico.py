from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import win32com.client
import os

import os
DB_SERVER = os.getenv('DB_SERVER', 'localhost')
DB_NAME = os.getenv('DB_NAME', 'DBBI')
DB_USER = os.getenv('DB_USER', 'sa')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_TABLE = 'CierreSucursales4'


def get_data_from_sql(departamento, ceco):
    SQLALCHEMY_DATABASE_URI = f"mssql+pyodbc://{DB_USER}:{DB_PASSWORD}@{DB_SERVER}/{DB_NAME}?driver=ODBC+Driver+17+for+SQL+Server"
    engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=False)
    query = text(f"""
    SELECT FORMAT(SUM(CAST([Val. Cont.] AS DECIMAL(18,2))), 'C', 'es-MX') AS Total
    FROM {DB_TABLE}
    WHERE [Ceco]=:ceco AND [Departamento]=:departamento AND [Accion]='BAJA';
    """)
    with engine.connect() as connection:
        result = connection.execute(query, {"ceco": ceco, "departamento": departamento})
        return result.fetchall()


def actualizar_excel_y_generar_pdf(ruta_excel, output_dir, departamento, ceco):
    ruta_temp = None
    excel = None
    try:
        datos = get_data_from_sql(departamento, ceco)
        if not datos:
            return
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        ruta_pdf = os.path.join(output_dir, f"Informe Tecnico_{departamento}_{ceco}.pdf")
        ruta_temp = ruta_excel.replace('.xlsx', '_temp.xlsx')
        wb = load_workbook(filename=ruta_excel)
        if 'Tecnico' not in wb.sheetnames:
            return
        hoja_baja = wb['Tecnico']
        for i, row in enumerate(datos):
            fila_actual = 14 + i
            hoja_baja[f'H{fila_actual}'] = row[0]
        wb.save(ruta_temp)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_temp))
        hoja_baja = wb.Sheets("Tecnico")
        hoja_baja.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None
    finally:
        if ruta_temp and os.path.exists(ruta_temp):
            try:
                os.remove(ruta_temp)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
